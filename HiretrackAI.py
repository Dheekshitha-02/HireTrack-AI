import os
import base64
import pandas as pd
import re
import spacy
from datetime import datetime, timedelta
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from bs4 import BeautifulSoup
from email.utils import parsedate_to_datetime
from pytz import timezone
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill
import schedule
import time


# Setup
SCOPES = ['https://www.googleapis.com/auth/gmail.readonly']
nlp = spacy.load("en_core_web_sm")
is_initial_run = True  # Flag to track first execution


def authenticate_gmail():
    creds = None
    if os.path.exists('token.json'):
        creds = Credentials.from_authorized_user_file('token.json', SCOPES)
    if not creds or not creds.valid:
        flow = InstalledAppFlow.from_client_secrets_file('credentials.json', SCOPES)
        flow.redirect_uri = "urn:ietf:wg:oauth:2.0:oob"
        auth_url, _ = flow.authorization_url(prompt='consent')
        print("\nOpen this URL in your browser and authorize access:")
        print(auth_url)
        code = input("\nPaste the authorization code here: ")
        flow.fetch_token(code=code)
        creds = flow.credentials
        with open('token.json', 'w') as token_file:
            token_file.write(creds.to_json())
    return build('gmail', 'v1', credentials=creds)

def extract_email_details(service):
    if is_initial_run:
        query = "(subject:application OR subject:rejected OR subject:unfortunately) newer_than:730d"
    else:
        query = "(subject:application OR subject:rejected OR subject:unfortunately) newer_than:30m"
    all_messages = []
    next_page_token = None

    while True:
        response = service.users().messages().list(
            userId='me',
            q=query,
            maxResults=100,
            pageToken=next_page_token
        ).execute()
        all_messages.extend(response.get('messages', []))
        next_page_token = response.get('nextPageToken')
        if not next_page_token:
            break

    data = []
    now = datetime.now(timezone('US/Eastern'))
    threshold = now - (timedelta(days=730) if is_initial_run else timedelta(minutes=30))
    for msg in all_messages:
        msg_data = service.users().messages().get(userId='me', id=msg['id'], format='full').execute()
        headers = msg_data['payload']['headers']
        subject = next((h['value'] for h in headers if h['name'] == 'Subject'), '')
        from_email = next((h['value'] for h in headers if h['name'].lower() == 'from'), '')

        subject_lower = subject.lower()
        from_email_lower = from_email.lower()

        if any(kw in subject_lower for kw in ["job alert", "new jobs", "recommended jobs", "job opportunities", "jobs for you"]):
            continue
        if "check out the status" in subject_lower and "jobs-noreply@linkedin.com" in from_email_lower:
            continue

        date_raw = next((h['value'] for h in headers if h['name'] == 'Date'), '')
        try:
            dt_obj = parsedate_to_datetime(date_raw).astimezone(timezone('US/Eastern'))
            if dt_obj < threshold:
                continue  # skip older emails
            date = dt_obj.date().isoformat()
            time = dt_obj.strftime("%H:%M")
        except Exception:
            date = ''
            time = ''

        body = extract_full_body_text(msg_data['payload']) or ''
        full_text = f"{subject} {body}"

        if "jobs-noreply@linkedin.com" in from_email_lower and "check out the status of your applications" in body.lower():
            continue
        if "jobs-noreply@linkedin.com" in from_email_lower and "your application was viewed" in subject_lower:
            continue
        if any(domain in from_email_lower for domain in ["jobalerts-noreply@linkedin.com", "editors-noreply@linkedin.com"]):
            continue

        status, trigger = classify_status_with_phrase(subject, body)
        if status not in ["Applied", "Interview", "Rejected"]:
            continue

        if "jobs-noreply@linkedin.com" in from_email_lower and "your application was sent to" in subject_lower:
            company = subject.split(" to ")[-1].strip().title()

            role_lines = [line.strip() for line in body.splitlines() if 5 < len(line.strip()) < 100]
            role = "Unknown"

            for idx, line in enumerate(role_lines):
                if "your application was sent to" in line.lower():
                    for lookahead in range(idx + 1, len(role_lines)):
                        next_line = role_lines[lookahead]
                        if any(k in next_line.lower() for k in ["engineer", "developer", "scientist", "analyst","programmer", "consultant", "intern", "manager"]):
                            role = next_line.strip().title()
                            break
                    break

        if "jobs-noreply@linkedin.com" in from_email_lower and subject_lower.startswith("your application to"):
            role_match = re.search(r'application to (.+?) at [A-Z][\w &().-]+', full_text, re.IGNORECASE)
            company_match = re.search(r'application to .+? at ([A-Z][\w &().-]+)', full_text, re.IGNORECASE)
            role = role_match.group(1).strip().title() if role_match else extract_role_spacy(full_text)
            company = company_match.group(1).strip().title() if company_match else extract_company_spacy(full_text)

        else:
            company = extract_company_spacy(full_text)
            role = extract_role_spacy(full_text)

        data.append({
            'Company': company,
            'Job Role': role,
            'Status': status,
            'Classification Phrase': trigger,
            'Date Applied': date,
            'Time Received': time
        })
    return data

def extract_full_body_text(payload):
    texts = []
    def recurse(part):
        if 'parts' in part:
            for subpart in part['parts']:
                recurse(subpart)
        elif 'body' in part and 'data' in part['body']:
            try:
                data = base64.urlsafe_b64decode(part['body']['data'])
                soup = BeautifulSoup(data, 'html.parser')
                text = soup.get_text(separator=' ', strip=True)
                if text:
                    texts.append(text)
            except Exception:
                pass
    recurse(payload)
    return ' '.join(texts)

def classify_status_with_phrase(subject, body):
    subject_clean = subject.lower().strip()
    body_clean = body.lower().strip()
    full_text = f"{subject_clean} {body_clean}"

    rejection_keywords = [
        "unfortunately", "we regret", "not selected", "declined",
        "decided to move forward", "moved forward with other",
        "we will not be moving forward", "pursue other candidates", "not currently aligned with our needs",
        "do not see a strong match for your experience"
    ]
    interview_keywords = [
        "interview scheduled", "interview invite", "speak with you", "we would like to schedule"
    ]
    applied_keywords = [
        "your application to", "application was sent", "you applied for",
        "recruiting team will contact you", "has been received", "have received your application",
        "reviewing all applications", "successfully submitted your application", "will review your submission"
    ]

    if subject_clean.startswith("your update from"):
        for phrase in rejection_keywords:
            if phrase in body_clean:
                return "Rejected", phrase

    for phrase in rejection_keywords:
        if phrase in full_text:
            return "Rejected", phrase
    for phrase in interview_keywords:
        if phrase in full_text:
            return "Interview", phrase
    for phrase in applied_keywords:
        if phrase in full_text:
            return "Applied", phrase

    return "Applied", "unknown"

def extract_company_spacy(text):
    blacklist = {"linkedin", "gmail", "workday", "ashbyhq", "greenhouse", "jobvite", "icims"}
    match = re.search(r'your update from\s+(.+?)(?:\.|\n|$)', text, re.IGNORECASE)
    if match:
        name = match.group(1).strip().title()
        if name.lower() not in blacklist and len(name.split()) <= 6:
            return name

    doc = nlp(text)
    signature_match = re.search(r'(Best Regards,|Regards,|Thanks,)\s*([A-Za-z &]+)', text, re.IGNORECASE)
    if signature_match:
        possible_name = signature_match.group(2).strip().title()
        if possible_name.lower() not in blacklist and len(possible_name.split()) <= 4:
            return possible_name

    for ent in doc.ents:
        if ent.label_ == "ORG":
            name = ent.text.strip().title()
            if name.lower() not in blacklist and len(name.split()) <= 4:
                return name
    return "Unknown"

def extract_role_spacy(text):
    match = re.search(r'application to (.+?) at [A-Z][\w &().-]+', text, re.IGNORECASE)
    if match:
        return match.group(1).strip().title()

    doc = nlp(text)
    keywords = ["engineer", "scientist","programmer", "developer", "analyst", "consultant", "intern", "manager", "specialist", "researcher"]
    for ent in doc.ents:
        if ent.label_ in ["WORK_OF_ART", "PRODUCT", "ORG", "JOB_TITLE", "NORP"]:
            if any(k in ent.text.lower() for k in keywords):
                title = ent.text.strip().title()
                return ' '.join(title.split()[:4])
    return "Unknown"

def save_to_excel(data):
    if not data:
        print("\n⚠ No relevant job application emails found.")
        return

    new_df = pd.DataFrame(data)
    new_df = new_df[['Company', 'Job Role', 'Status', 'Classification Phrase', 'Date Applied', 'Time Received']]
    new_df.drop_duplicates(inplace=True)

    file_name = "job_applications_hireai.xlsx"

    if os.path.exists(file_name):
        existing_df = pd.read_excel(file_name)
        combined_df = pd.concat([existing_df, new_df], ignore_index=True)
        combined_df.drop_duplicates(subset=['Company', 'Job Role', 'Date Applied', 'Time Received'], inplace=True)
    else:
        combined_df = new_df

    combined_df.to_excel(file_name, index=False)

    wb = load_workbook(file_name)
    ws = wb.active

    dv = DataValidation(type="list", formula1='"Applied,Interview,Rejected,Offer"', allow_blank=True)
    ws.add_data_validation(dv)

    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')   # Applied
    yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')  # Interview
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')     # Rejected
    blue_fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')    # Offer

    for row in range(2, ws.max_row + 1):
        cell = ws[f'C{row}']
        dv.add(cell)
        status = str(cell.value).strip().lower()
        if "applied" in status:
            cell.fill = green_fill
        elif "interview" in status:
            cell.fill = yellow_fill
        elif "rejected" in status:
            cell.fill = red_fill
        elif "offer" in status:
            cell.fill = blue_fill

    combined_df['Date Applied'] = pd.to_datetime(combined_df['Date Applied'], errors='coerce')
    combined_df['Time Received'] = pd.to_datetime(combined_df['Time Received'], format="%H:%M", errors='coerce').dt.time

    combined_df = combined_df.sort_values(by=["Date Applied", "Time Received"], ascending=[False, False])

    combined_df.drop_duplicates(subset=['Company', 'Job Role', 'Date Applied', 'Time Received'], inplace=True)

    wb.save(file_name)
    print("\nExcel updated:", file_name)

def main():
    print("Authenticating...")
    service = authenticate_gmail()
    print("Fetching emails...")
    data = extract_email_details(service)
    print(f"Extracted {len(data)} records.")
    if data:
        print("Saving to Excel...")
    save_to_excel(data)
    print("Done.")

def main():
    print("Authenticating...")
    service = authenticate_gmail()
    print("Fetching emails...")
    data = extract_email_details(service)
    print(f"Extracted {len(data)} records.")
    if data:
        print("Saving to Excel...")
    save_to_excel(data)
    print("Done.")

def run_tracker():
    global is_initial_run
    print("\nRunning job...")
    service = authenticate_gmail()
    data = extract_email_details(service)
    print(f"Extracted {len(data)} records.")
    if data:
        print("Saving to Excel...")
        save_to_excel(data)
    else:
        print("No new relevant emails found.")
    print("Done.")
    is_initial_run = False  # Mark that first run is completed


if __name__ == '__main__':
    run_tracker()  # Initial manual run
    schedule.every(30).minutes.do(run_tracker)
    print("Job tracker is scheduled to run every 30 minutes...")
    while True:
        schedule.run_pending()
        time.sleep(1)